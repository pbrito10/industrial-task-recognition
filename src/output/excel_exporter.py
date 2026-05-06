from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill

from src.output.metrics_snapshot import MetricsSnapshot
from src.output.output_interface import OutputInterface
from src.tracking.cycle_result import CycleResult
from src.tracking.order_matching import diagnose_order
from src.tracking.task_event import TaskEvent

# Cor de destaque para a zona gargalo (amarelo-âmbar)
_BOTTLENECK_FILL = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
_HEADER_FONT     = Font(bold=True)
_TASK_COLUMN     = "Tarefa/Zona"
_CYCLE_COLUMNS   = [
    "Nº Ciclo",
    "Início",
    "Fim",
    "Duração (s)",
    "Resultado do sistema",
    "Sequência registada",
    "Problema detetado",
    "Classificação manual",
    "Observações",
]

# Mapeamento direto para evitar ternário aninhado em _write_events
_FORCED_LABEL: dict[bool, str] = {True: "Sim", False: "Não"}


@dataclass(frozen=True)
class _EventExportRow:
    event: TaskEvent
    counts_as_interruption: bool = False


def _cycle_diagnosis(cycle_result: CycleResult):
    return diagnose_order(cycle_result.actual_sequence, cycle_result.expected_sequence)


class ExcelExporter(OutputInterface):
    """Exporta os dados da sessão para um ficheiro .xlsx no fim da sessão.

    Gera quatro folhas: Resumo, Métricas por Zona, Ciclos e Eventos.
    A tarefa/zona gargalo é destacada a amarelo na folha de métricas.
    Na folha Eventos, a coluna Tipo separa tempo produtivo de interrupções
    sem confundir interrupções de análise com timeouts forçados.

    write() é chamado uma vez no fim — não é para uso online como o DashboardWriter.
    """

    def __init__(self, output_dir: Path, session_start: datetime) -> None:
        self._output_dir    = output_dir
        self._session_start = session_start
        self._events:         list[_EventExportRow] = []
        self._cycle_results:  dict[int, CycleResult] = {}

        output_dir.mkdir(parents=True, exist_ok=True)

    def add_event(self, event: TaskEvent, counts_as_interruption: bool = False) -> None:
        """Acumula TaskEvents durante a sessão para exportar no fim."""
        self._events.append(_EventExportRow(event, counts_as_interruption))

    def add_cycle_result(self, cycle_result: CycleResult) -> None:
        """Regista o ciclo fechado para exportação."""
        self._cycle_results[cycle_result.cycle_number] = cycle_result

    def write(self, snapshot: MetricsSnapshot) -> None:
        """Gera o ficheiro Excel com todas as folhas."""
        filename = f"sessao_{self._session_start.strftime('%Y-%m-%d_%Hh%M')}.xlsx"
        path     = self._output_dir / filename

        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            self._write_summary(writer, snapshot)
            self._write_zone_metrics(writer, snapshot)
            self._write_cycles(writer)
            self._write_events(writer)

    # ── Folhas ───────────────────────────────────────────────────────────────

    def _write_summary(self, writer: pd.ExcelWriter, snapshot: MetricsSnapshot) -> None:
        """Folha 'Resumo': métricas globais da sessão numa única tabela de dois campos."""
        cycle   = snapshot.cycle_metrics
        avg_s   = "—"
        std_s   = "—"
        if cycle.count():
            avg_s = round(cycle.average().total_seconds(), 2)
            std_s = round(cycle.std_deviation().total_seconds(), 2)

        rows = [
            ("Data",                    self._session_start.strftime("%Y-%m-%d %H:%M")),
            ("Duração total (s)",       round(snapshot.session_duration.total_seconds(), 1)),
            ("Ciclos completos",        cycle.count()),
            ("Tempo médio ciclo (s)",   avg_s),
            ("Desvio padrão ciclo (s)", std_s),
            ("% Tempo produtivo",       round(snapshot.productive_percentage, 1)),
            ("% Tempo transição",       round(snapshot.transition_percentage, 1)),
            ("% Tempo interrupção",     round(snapshot.interruption_percentage, 1)),
            ("Tarefa/Zona gargalo",     snapshot.bottleneck_zone or "—"),
        ]

        df = pd.DataFrame(rows, columns=["Métrica", "Valor"])
        df.to_excel(writer, sheet_name="Resumo", index=False)
        self._bold_headers(writer, "Resumo", df)

    def _write_zone_metrics(self, writer: pd.ExcelWriter, snapshot: MetricsSnapshot) -> None:
        """Folha 'Métricas por Zona': estatísticas por tarefa/zona com gargalo destacado."""
        rows = []
        for zone_name, metrics in snapshot.task_metrics.items():
            if metrics.count() == 0:
                continue
            rows.append({
                _TASK_COLUMN:      zone_name,
                "Mínimo (s)":      round(metrics.minimum().total_seconds(), 3),
                "Médio (s)":       round(metrics.average().total_seconds(), 3),
                "Máximo (s)":      round(metrics.maximum().total_seconds(), 3),
                "Desvio Padrão (s)": round(metrics.std_deviation().total_seconds(), 3),
                "Ocorrências":     metrics.count(),
            })

        df = pd.DataFrame(rows)
        df.to_excel(writer, sheet_name="Métricas por Zona", index=False)
        self._bold_headers(writer, "Métricas por Zona", df)
        self._highlight_bottleneck(writer, "Métricas por Zona", df, snapshot.bottleneck_zone)

    def _write_cycles(self, writer: pd.ExcelWriter) -> None:
        """Folha 'Ciclos': uma linha por ciclo fechado."""
        rows = []
        for cycle_result in sorted(self._cycle_results.values(), key=lambda cycle: cycle.cycle_number):
            diagnosis = _cycle_diagnosis(cycle_result)
            rows.append({
                "Nº Ciclo":             cycle_result.cycle_number,
                "Início":               cycle_result.start_time.strftime("%H:%M:%S"),
                "Fim":                  cycle_result.end_time.strftime("%H:%M:%S"),
                "Duração (s)":          round(cycle_result.duration.total_seconds(), 2),
                "Resultado do sistema": diagnosis.result,
                "Sequência registada": " → ".join(cycle_result.actual_sequence),
                "Problema detetado":    diagnosis.problem,
                "Classificação manual": "",
                "Observações":          "",
            })

        df = pd.DataFrame(rows, columns=_CYCLE_COLUMNS)
        df.to_excel(writer, sheet_name="Ciclos", index=False)
        self._bold_headers(writer, "Ciclos", df)

    def _write_events(self, writer: pd.ExcelWriter) -> None:
        """Folha 'Eventos': uma linha por TaskEvent com categoria produtivo/interrupção."""
        rows = [
            {
                "Ciclo":       row.event.cycle_number,
                "Tarefa/Zona": row.event.zone_name,
                "Tipo":        self._event_type_label(row),
                "Início":      row.event.start_time.strftime("%H:%M:%S.%f")[:-3],
                "Fim":         row.event.end_time.strftime("%H:%M:%S.%f")[:-3],
                "Duração (s)": round(row.event.duration.total_seconds(), 3),
                "Forçado":     _FORCED_LABEL[row.event.was_forced],
            }
            for row in self._events
        ]

        df = pd.DataFrame(rows)
        df.to_excel(writer, sheet_name="Eventos", index=False)
        self._bold_headers(writer, "Eventos", df)

    def _event_type_label(self, row: _EventExportRow) -> str:
        if row.counts_as_interruption or row.event.was_forced:
            return "Interrupção"
        return "Produtivo"

    # ── Formatação ────────────────────────────────────────────────────────────

    def _bold_headers(self, writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame) -> None:
        """Aplica bold à linha de cabeçalho da folha indicada."""
        sheet = writer.sheets[sheet_name]
        for col_idx in range(1, len(df.columns) + 1):
            sheet.cell(row=1, column=col_idx).font = _HEADER_FONT

    def _highlight_bottleneck(
        self,
        writer: pd.ExcelWriter,
        sheet_name: str,
        df: pd.DataFrame,
        bottleneck: str | None,
    ) -> None:
        """Destaca a amarelo a linha correspondente à tarefa/zona gargalo."""
        if bottleneck is None or _TASK_COLUMN not in df.columns:
            return

        sheet    = writer.sheets[sheet_name]
        num_cols = len(df.columns)

        # Linha 1 é o cabeçalho — dados começam na linha 2
        for row_idx, zone_name in enumerate(df[_TASK_COLUMN], start=2):
            if zone_name == bottleneck:
                for col_idx in range(1, num_cols + 1):
                    sheet.cell(row=row_idx, column=col_idx).fill = _BOTTLENECK_FILL
