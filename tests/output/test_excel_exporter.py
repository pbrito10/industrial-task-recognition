from datetime import datetime, timedelta

import openpyxl
import pytest

from src.metrics.cycle_metrics import CycleMetrics
from src.metrics.task_metrics import TaskMetrics
from src.output.excel_exporter import ExcelExporter
from src.output.metrics_snapshot import MetricsSnapshot
from src.tracking.cycle_result import CycleResult
from src.tracking.task_event import TaskEvent

_SESSION_START = datetime(2024, 3, 15, 9, 0, 0)
_T0 = _SESSION_START


def _event(zone="Porca", duration_s=5.0, forced=False, cycle=1) -> TaskEvent:
    start = _T0
    end   = _T0 + timedelta(seconds=duration_s)
    return TaskEvent.create(zone, start, end, cycle_number=cycle, was_forced=forced)


def _cycle(duration_s=60.0, in_order=True, number=1) -> CycleResult:
    return CycleResult(
        duration=timedelta(seconds=duration_s),
        cycle_number=number,
        sequence_in_order=in_order,
        actual_sequence=["Porca", "Saida"],
    )


def _snapshot(bottleneck=None) -> MetricsSnapshot:
    tm = TaskMetrics()
    tm.add(timedelta(seconds=5))
    tm.add(timedelta(seconds=7))
    cm = CycleMetrics()
    cm.add(timedelta(seconds=60), True)
    return MetricsSnapshot(
        task_metrics={"Porca": tm},
        cycle_metrics=cm,
        productive_time=timedelta(seconds=12),
        transition_time=timedelta(seconds=5),
        interruption_time=timedelta(seconds=3),
        productive_percentage=60.0,
        transition_percentage=25.0,
        interruption_percentage=15.0,
        bottleneck_zone=bottleneck,
        session_duration=timedelta(seconds=90),
        captured_at=_T0,
    )


@pytest.fixture
def exporter(tmp_path):
    return ExcelExporter(tmp_path, _SESSION_START)


def _open_workbook(tmp_path):
    files = list(tmp_path.glob("sessao_*.xlsx"))
    assert len(files) == 1
    return openpyxl.load_workbook(str(files[0]))


# --- Criação do ficheiro ---

def test_creates_xlsx_file(exporter, tmp_path):
    exporter.write(_snapshot())
    assert len(list(tmp_path.glob("sessao_*.xlsx"))) == 1


def test_filename_contains_session_date(exporter, tmp_path):
    exporter.write(_snapshot())
    files = list(tmp_path.glob("sessao_*.xlsx"))
    assert "2024-03-15" in files[0].name


# --- Quatro folhas ---

def test_has_four_sheets(exporter, tmp_path):
    exporter.write(_snapshot())
    wb = _open_workbook(tmp_path)
    assert set(wb.sheetnames) == {"Resumo", "Métricas por Zona", "Ciclos", "Eventos"}


# --- Folha Resumo ---

def test_resumo_has_data_rows(exporter, tmp_path):
    exporter.write(_snapshot())
    wb  = _open_workbook(tmp_path)
    ws  = wb["Resumo"]
    # linha 1 é header, deve haver linhas de dados
    assert ws.max_row > 1


def test_resumo_bottleneck_value(exporter, tmp_path):
    exporter.write(_snapshot(bottleneck="Porca"))
    wb   = _open_workbook(tmp_path)
    ws   = wb["Resumo"]
    values = [ws.cell(row=r, column=2).value for r in range(2, ws.max_row + 1)]
    assert "Porca" in values


def test_resumo_no_bottleneck_shows_dash(exporter, tmp_path):
    exporter.write(_snapshot(bottleneck=None))
    wb     = _open_workbook(tmp_path)
    ws     = wb["Resumo"]
    values = [ws.cell(row=r, column=2).value for r in range(2, ws.max_row + 1)]
    assert "—" in values


# --- Folha Métricas por Zona ---

def test_zone_metrics_has_porca_row(exporter, tmp_path):
    exporter.write(_snapshot())
    wb  = _open_workbook(tmp_path)
    ws  = wb["Métricas por Zona"]
    # linha 2 = primeira linha de dados
    assert ws.cell(row=2, column=1).value == "Porca"


def test_zone_metrics_bottleneck_highlighted(exporter, tmp_path):
    exporter.write(_snapshot(bottleneck="Porca"))
    wb   = _open_workbook(tmp_path)
    ws   = wb["Métricas por Zona"]
    fill = ws.cell(row=2, column=1).fill
    # openpyxl devolve ARGB (8 chars): "00FFD966" — verificamos os últimos 6
    assert fill.fgColor.rgb.endswith("FFD966")


def test_zone_metrics_no_bottleneck_not_highlighted(exporter, tmp_path):
    exporter.write(_snapshot(bottleneck=None))
    wb   = _open_workbook(tmp_path)
    ws   = wb["Métricas por Zona"]
    fill = ws.cell(row=2, column=1).fill
    # Sem bottleneck, a célula não tem fill amarelo
    assert fill.fgColor.rgb != "FFD966"


# --- Folha Ciclos ---

def test_cycles_written_after_add_cycle(exporter, tmp_path):
    exporter.add_event(_event(cycle=1))
    exporter.add_cycle_result(_cycle(number=1))
    exporter.write(_snapshot())
    wb = _open_workbook(tmp_path)
    ws = wb["Ciclos"]
    assert ws.max_row >= 2  # header + 1 ciclo


def test_cycles_sequence_in_order_label(exporter, tmp_path):
    exporter.add_event(_event(cycle=1))
    exporter.add_cycle_result(_cycle(in_order=True, number=1))
    exporter.write(_snapshot())
    wb     = _open_workbook(tmp_path)
    ws     = wb["Ciclos"]
    values = [ws.cell(row=r, column=5).value for r in range(2, ws.max_row + 1)]
    assert "Sim" in values


def test_cycles_out_of_order_label(exporter, tmp_path):
    exporter.add_event(_event(cycle=1))
    exporter.add_cycle_result(_cycle(in_order=False, number=1))
    exporter.write(_snapshot())
    wb     = _open_workbook(tmp_path)
    ws     = wb["Ciclos"]
    values = [ws.cell(row=r, column=5).value for r in range(2, ws.max_row + 1)]
    assert "Não" in values


# --- Folha Eventos ---

def test_events_written_after_add_event(exporter, tmp_path):
    exporter.add_event(_event("Porca"))
    exporter.add_event(_event("Montagem"))
    exporter.write(_snapshot())
    wb = _open_workbook(tmp_path)
    ws = wb["Eventos"]
    assert ws.max_row == 3  # header + 2 eventos


def test_events_forced_label(exporter, tmp_path):
    exporter.add_event(_event("Porca", forced=True))
    exporter.write(_snapshot())
    wb     = _open_workbook(tmp_path)
    ws     = wb["Eventos"]
    values = [ws.cell(row=r, column=6).value for r in range(2, ws.max_row + 1)]
    assert "Sim" in values


def test_events_not_forced_label(exporter, tmp_path):
    exporter.add_event(_event("Porca", forced=False))
    exporter.write(_snapshot())
    wb     = _open_workbook(tmp_path)
    ws     = wb["Eventos"]
    values = [ws.cell(row=r, column=6).value for r in range(2, ws.max_row + 1)]
    assert "Não" in values


# --- Headers a bold ---

def test_headers_are_bold(exporter, tmp_path):
    # Ciclos e Eventos só têm colunas se tiverem dados — adicionar um registo de cada
    exporter.add_event(_event(cycle=1))
    exporter.add_cycle_result(_cycle(number=1))
    exporter.write(_snapshot())
    wb = _open_workbook(tmp_path)
    for sheet_name in ["Resumo", "Métricas por Zona", "Ciclos", "Eventos"]:
        ws   = wb[sheet_name]
        cell = ws.cell(row=1, column=1)
        assert cell.font.bold, f"Header de '{sheet_name}' não está a bold"


# --- Sem dados não rebenta ---

def test_write_with_no_events_no_crash(exporter, tmp_path):
    exporter.write(_snapshot())
    wb = _open_workbook(tmp_path)
    assert "Eventos" in wb.sheetnames
